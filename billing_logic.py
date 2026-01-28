# billing_logic.py
import os
import re
import fitz  # PyMuPDF
import tempfile
import shutil
from docx2pdf import convert
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


# -------------------------------------------------
# DOCX PAGE COUNT
# -------------------------------------------------
def get_docx_page_count(docx_path):
    try:
        temp_dir = tempfile.mkdtemp()
        temp_pdf = os.path.join(temp_dir, "temp.pdf")
        convert(docx_path, temp_pdf)
        pages = len(PdfReader(temp_pdf).pages)
        shutil.rmtree(temp_dir)
        return pages
    except Exception:
        return "NA"


# -------------------------------------------------
# FIND ORIGINAL FOLDER
# -------------------------------------------------
def find_original_folder(job_path):
    for name in os.listdir(job_path):
        full = os.path.join(job_path, name)
        if os.path.isdir(full) and name.lower().replace(" ", "").endswith("original"):
            return full
    return None


# -------------------------------------------------
# FIND LATEST FINAL DRAFT PDF PAGES
# -------------------------------------------------
def get_latest_final_draft_pages(job_path):
    final_folder = None

    for name in os.listdir(job_path):
        if name.lower().startswith("6.final"):
            final_folder = os.path.join(job_path, name)
            break

    if not final_folder or not os.path.isdir(final_folder):
        return ("NA", "NA")

    drafts = []

    for file in os.listdir(final_folder):
        if not file.lower().endswith(".pdf"):
            continue

        match = re.search(r"_draft\s*(\d+)([a-z]?)", file.lower())
        if match:
            draft_num = int(match.group(1))
            suffix = match.group(2) or ""
            drafts.append((draft_num, suffix, file))

    if not drafts:
        return ("NA", "NA")

    # 🔥 sort by number, then suffix
    drafts.sort(key=lambda x: (x[0], x[1]))

    latest_num, latest_suffix, latest_file = drafts[-1]

    try:
        doc = fitz.open(os.path.join(final_folder, latest_file))
        return (latest_file, len(doc))
    except Exception:
        return (latest_file, "NA")

# -------------------------------------------------
# PROCESS FILES (RECURSIVE)
# -------------------------------------------------
def process_folder(folder_path, job_name, edit_folder):
    summary_rows = []
    detail_rows = []

    for root, _, files in os.walk(folder_path):
        inside_folder = os.path.relpath(root, folder_path)
        if inside_folder == ".":
            inside_folder = "ROOT"

        for file in files:
            full = os.path.join(root, file)

            # ---------- PDF ----------
            if file.lower().endswith(".pdf"):
                doc = fitz.open(full)
                total_comments = 0
                pages_with_comments = set()

                for page_index in range(len(doc)):
                    annots = doc[page_index].annots()
                    if annots:
                        for annot in annots:
                            total_comments += 1
                            pages_with_comments.add(page_index + 1)

                            detail_rows.append({
                                "Job Name": job_name,
                                "Edit Folder": edit_folder,
                                "Inside Folder": inside_folder,
                                "File Name": file,
                                "File Type": "PDF",
                                "Page Number": page_index + 1,
                                "Comment Type": annot.type[1] if annot.type else "Unknown",
                                "Author": annot.info.get("title", ""),
                                "Comment Text": annot.info.get("content", "").strip()
                            })

                summary_rows.append({
                    "Job Name": job_name,
                    "Edit Folder": edit_folder,
                    "Inside Folder": inside_folder,
                    "File Name": file,
                    "File Type": "PDF",
                    "Total Pages": len(doc),
                    "Pages with Comments": len(pages_with_comments),
                    "Total Comments": total_comments
                })

            # ---------- DOCX ----------
            elif file.lower().endswith(".docx"):
                pages = get_docx_page_count(full)
                summary_rows.append({
                    "Job Name": job_name,
                    "Edit Folder": edit_folder,
                    "Inside Folder": inside_folder,
                    "File Name": file,
                    "File Type": "DOCX",
                    "Total Pages": pages,
                    "Pages with Comments": "NA",
                    "Total Comments": "NA"
                })

    return summary_rows, detail_rows


# -------------------------------------------------
# SINGLE MODE
# -------------------------------------------------
def run_single_mode(folder_path):
    return process_folder(folder_path, "SingleJob", "NA")


# -------------------------------------------------
# BATCH MODE
# -------------------------------------------------
def run_batch_mode(root_folder):
    all_summary = []
    all_details = []
    final_draft_pages = {}

    for job in os.listdir(root_folder):
        job_path = os.path.join(root_folder, job)
        if not os.path.isdir(job_path):
            continue

        final_draft_pages[job] = get_latest_final_draft_pages(job_path)


        original = find_original_folder(job_path)
        if not original:
            print(f"⚠️ Skipping {job} (Original folder not found)")
            continue

        for sub in os.listdir(original):
            if not sub.startswith("Edits_"):
                continue

            edits_path = os.path.join(original, sub)
            if not os.path.isdir(edits_path):
                continue

            print(f"✅ Processing: {job} → {sub}")

            summary, details = process_folder(edits_path, job, sub)
            all_summary.extend(summary)
            all_details.extend(details)

    return all_summary, all_details, final_draft_pages


# -------------------------------------------------
# EXCEL OUTPUT
# -------------------------------------------------
root_fill = PatternFill("solid", fgColor="D9EAF7")
subfolder_fill = PatternFill("solid", fgColor="E2F0D9")
header_fill = PatternFill("solid", fgColor="DDDDDD")
job_fill = PatternFill("solid", fgColor="FFFF00")
bold_font = Font(bold=True)


def generate_master_excel(summary_rows, detail_rows, final_draft_pages, output_folder):
    output = os.path.join(output_folder, "Master_Billing_Report.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    headers = [
        "Job Name",
        "Edit Folder",
        "Inside Folder",
        "File Name",
        "File Type",
        "Total Pages",
        "Pages with Comments",
        "Total Comments"
    ]
    ws.append(headers)

    for c in range(1, 9):
        ws.cell(row=1, column=c).fill = header_fill
        ws.cell(row=1, column=c).font = bold_font

    jobs = {}
    for row in summary_rows:
        jobs.setdefault(row["Job Name"], []).append(row)

    for job_name, rows in jobs.items():
        ws.append([f"JOB NAME ---- {job_name}"])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.cell(row=r, column=1).fill = job_fill
        ws.cell(row=r, column=1).font = bold_font

        total_docx_pages = 0
        total_pdf_pages = 0
        total_pdf_comments = 0

        for rdata in rows:
            ws.append([
                rdata["Job Name"],
                rdata["Edit Folder"],
                rdata["Inside Folder"],
                rdata["File Name"],
                rdata["File Type"],
                rdata["Total Pages"],
                rdata["Pages with Comments"],
                rdata["Total Comments"]
            ])

            row_idx = ws.max_row
            fill = root_fill if rdata["Inside Folder"] == "ROOT" else subfolder_fill
            for c in range(1, 9):
                ws.cell(row=row_idx, column=c).fill = fill

            if rdata["File Type"] == "DOCX" and rdata["Total Pages"] != "NA":
                total_docx_pages += int(rdata["Total Pages"])

            if rdata["File Type"] == "PDF":
                if rdata["Pages with Comments"] != "NA":
                    total_pdf_pages += int(rdata["Pages with Comments"])
                if rdata["Total Comments"] != "NA":
                    total_pdf_comments += int(rdata["Total Comments"])

        # TOTAL ROW
        ws.append([
            "Total",
            "",
            "",
            "",
            "DOCX pages",
            total_docx_pages,
            total_pdf_pages,
            total_pdf_comments
        ])

        tr = ws.max_row
        for c in range(1, 9):
            ws.cell(row=tr, column=c).font = bold_font

        # FINAL DRAFT ROW
        draft_name, draft_pages = final_draft_pages.get(job_name, ("NA", "NA"))

        ws.append([
            f"Final Draft Total Pages ({draft_name})",
            "",
            "",
            "",
            "",
            draft_pages,
            "",
            ""
        ])

        fr = ws.max_row
        for c in range(1, 9):
            ws.cell(row=fr, column=c).font = bold_font

        ws.append([])

    if detail_rows:
        ws2 = wb.create_sheet("PDF_Comments")
        ws2.append(list(detail_rows[0].keys()))
        for row in detail_rows:
            ws2.append(list(row.values()))

    wb.save(output)
    return output
